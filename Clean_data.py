"""
================================================================================
  BC VER1 BUILDER — Toàn bộ quy trình xử lý dữ liệu từ file gốc → ver1
  Author  : Antigravity AI Data Analyst
  Date    : 04/2026
  Input   : D:/ver1/data/ver1.xlsx by default
  Output  : D:/ver1/ver1_cleaned.xlsx by default

  Mô tả:
    Script này tái tạo toàn bộ pipeline xử lý dữ liệu từ file bc gốc,
    bao gồm:
      - Làm sạch dữ liệu (Raw Data)
      - Thêm cột phân loại Ngành Hàng
      - Tính lại các chỉ số tài chính
      - Giữ lại tất cả các sheet NGOẠI TRỪ các sheet Pivot Table
    (Product Pivot và Customer Pivot bị bỏ qua vì được tạo tự động trong Excel)
================================================================================
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import os

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
PROJECT_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE   = os.environ.get('CHURN_CLEAN_INPUT', os.path.join(PROJECT_DIR, 'data', 'ver1.xlsx'))
OUTPUT_DIR   = os.environ.get('CHURN_CLEAN_OUTPUT_DIR', os.path.join(PROJECT_DIR, 'outputs'))
OUTPUT_FILE  = os.environ.get('CHURN_CLEAN_OUTPUT', os.path.join(OUTPUT_DIR, 'ver1_cleaned.xlsx'))
RAW_SHEET    = os.environ.get('CHURN_CLEAN_SHEET', 'Data Model')

# Sheets cần COPY NGUYÊN (không phải Pivot)
COPY_SHEETS = [
    'Product Dashboard',
    'Customer Dashboard',
    RAW_SHEET,
    'Strategic Insights',
    'RFM Table',
    'Pivot Khách Hàng',
    'Segment Overview',
    'Log GD Ghi Giảm KM',
    '⚠️ Cần Review',
]

# Sheets cần BỎ QUA (Pivot Table — Excel tự quản lý)
SKIP_SHEETS = ['Product Pivot', 'Customer Pivot']

os.makedirs(OUTPUT_DIR, exist_ok=True)
print(f"Output folder: {OUTPUT_DIR}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: ĐỌC DỮ LIỆU GỐC
# ─────────────────────────────────────────────────────────────────────────────
print("\n[1] Loading raw data from source file...")
df_raw = pd.read_excel(INPUT_FILE, sheet_name=RAW_SHEET)
print(f"    → {len(df_raw)} rows × {len(df_raw.columns)} cols")
print(f"    → Columns: {df_raw.columns.tolist()}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: LÀM SẠCH DỮ LIỆU (Data Cleaning)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2] Cleaning data...")

# 2.1 Xóa dòng hoàn toàn trống
before = len(df_raw)
df_raw.dropna(how='all', inplace=True)
print(f"    → Dropped {before - len(df_raw)} empty rows")

# 2.2 Trim whitespace tất cả cột string
for col in df_raw.select_dtypes(include='object').columns:
    df_raw[col] = df_raw[col].str.strip()

# 2.3 Ép kiểu số cho các cột tài chính
NUMERIC_COLS = [
    'Tiền Nhập', 'SL Xuất KM', 'SL Xuất', 'Tiền Xuất',
    'Doanh Thu', 'Lợi Nhuận', 'Tổng số lượng bán', '% SL Khuyến mãi'
]
for col in NUMERIC_COLS:
    if col in df_raw.columns:
        df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce').fillna(0)
print(f"    → {len(NUMERIC_COLS)} numeric columns coerced")

# 2.4 Parse ngày tháng
if 'Ngày Xuất' in df_raw.columns:
    df_raw['Ngày Xuất'] = pd.to_datetime(df_raw['Ngày Xuất'], errors='coerce', dayfirst=True)
    print(f"    → 'Ngày Xuất' parsed. Range: {df_raw['Ngày Xuất'].min().date()} → {df_raw['Ngày Xuất'].max().date()}")

# 2.5 Đảm bảo cột Year / Month đồng bộ với Ngày Xuất
if 'Ngày Xuất' in df_raw.columns:
    df_raw['Year']  = df_raw.get('Year',  pd.Series()).fillna(df_raw['Ngày Xuất'].dt.year).astype('Int64')
    df_raw['Month'] = df_raw.get('Month', pd.Series()).fillna(df_raw['Ngày Xuất'].dt.month).astype('Int64')
    print(f"    → Year/Month synced: {sorted(df_raw['Month'].dropna().unique().tolist())}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: TẠO CỘT PHÁI SINH (Feature Engineering)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3] Feature engineering...")

# 3.1 Phân loại Ngành Hàng từ tiền tố Mã sản phẩm
# Quy tắc: lấy tất cả ký tự chữ cái liên tiếp đầu tiên của cột Mã
# rồi map sang tên ngành hàng.
CATEGORY_MAP = {
    # Khăn Giấy
    'KV':'Khăn Giấy', 'KR':'Khăn Giấy', 'KDL':'Khăn Giấy',
    'KH':'Khăn Giấy', 'KL':'Khăn Giấy', 'KB':'Khăn Giấy',
    'KKDN':'Khăn Giấy', 'KVK':'Khăn Giấy',
    # Khăn Ướt
    'KU':'Khăn Ướt',
    # Giấy Vệ Sinh
    'GC':'Giấy Vệ Sinh', 'HM':'Giấy Vệ Sinh',
    'HMK':'Giấy Vệ Sinh', 'HMH':'Giấy Vệ Sinh',
    # Bàn Chải & Răng Miệng
    'BC':'Bàn Chải & Răng Miệng', 'BCN':'Bàn Chải & Răng Miệng',
    'BCTE':'Bàn Chải & Răng Miệng', 'BCTEM':'Bàn Chải & Răng Miệng',
    'BCTEN':'Bàn Chải & Răng Miệng', 'BCA':'Bàn Chải & Răng Miệng',
    'BCX':'Bàn Chải & Răng Miệng', 'BCOCEAN':'Bàn Chải & Răng Miệng',
    'BCHARU':'Bàn Chải & Răng Miệng', 'BCAKIRA':'Bàn Chải & Răng Miệng',
    'BCTEPANDA':'Bàn Chải & Răng Miệng', 'BCTEELEPHAN':'Bàn Chải & Răng Miệng',
    'BCUNIQUE':'Bàn Chải & Răng Miệng', 'BCSILISON':'Bàn Chải & Răng Miệng',
    'BCJEWELL':'Bàn Chải & Răng Miệng', 'BCTENEKO':'Bàn Chải & Răng Miệng',
    'KDR':'Bàn Chải & Răng Miệng', 'HTKDR':'Bàn Chải & Răng Miệng',
    'DCKAKA':'Bàn Chải & Răng Miệng', 'DCTD':'Bàn Chải & Răng Miệng',
    'LLDORCOTITAN':'Bàn Chải & Răng Miệng',
    # Tăm Bông
    'TBTE':'Tăm Bông', 'TB':'Tăm Bông', 'BTT':'Tăm Bông', 'BT':'Tăm Bông',
    # Thực Phẩm & Bánh Kẹo
    'KTN':'Thực Phẩm & Bánh Kẹo', 'HTK':'Thực Phẩm & Bánh Kẹo',
    'A':'Thực Phẩm & Bánh Kẹo', 'CC':'Thực Phẩm & Bánh Kẹo',
    'VMS':'Thực Phẩm & Bánh Kẹo', 'HTKD':'Thực Phẩm & Bánh Kẹo',
    'HTKSB':'Thực Phẩm & Bánh Kẹo', 'SD':'Thực Phẩm & Bánh Kẹo',
    'KGC':'Thực Phẩm & Bánh Kẹo', 'KGLC':'Thực Phẩm & Bánh Kẹo',
    'XX':'Thực Phẩm & Bánh Kẹo', 'CG':'Thực Phẩm & Bánh Kẹo',
    # Bánh Snack
    'HTBTE':'Bánh Snack', 'HTBSN':'Bánh Snack', 'HTBD':'Bánh Snack',
    'HTBX':'Bánh Snack', 'HTBG':'Bánh Snack', 'HTBXG':'Bánh Snack',
    'BWW':'Bánh Snack', 'BSOGI':'Bánh Snack', 'BBL':'Bánh Snack',
    # Nước Uống
    'NSSK':'Nước Uống', 'TBND':'Nước Uống', 'TBME':'Nước Uống',
    'NLREDBULL':'Nước Uống', 'NLSDN':'Nước Uống',
    # Chăm Sóc Nhà Cửa
    'NG':'Chăm Sóc Nhà Cửa', 'NGT':'Chăm Sóc Nhà Cửa',
    'JA':'Chăm Sóc Nhà Cửa', 'TC':'Chăm Sóc Nhà Cửa',
    'TNS':'Chăm Sóc Nhà Cửa', 'TLM':'Chăm Sóc Nhà Cửa',
    'BG':'Chăm Sóc Nhà Cửa', 'NRC':'Chăm Sóc Nhà Cửa',
    'HTF':'Chăm Sóc Nhà Cửa', 'HTD':'Chăm Sóc Nhà Cửa',
    'HTCTBC':'Chăm Sóc Nhà Cửa', 'LK':'Chăm Sóc Nhà Cửa',
    'SRT':'Chăm Sóc Nhà Cửa', 'XM':'Chăm Sóc Nhà Cửa',
    'HTNM':'Chăm Sóc Nhà Cửa', 'T':'Chăm Sóc Nhà Cửa',
    # Yến Sào & Sức Khỏe
    'BV':'Yến Sào & Sức Khỏe', 'CV':'Yến Sào & Sức Khỏe',
    'SBOMGROW':'Yến Sào & Sức Khỏe',
    # Mẹ & Bé
    'NV':'Mẹ & Bé', 'KT':'Mẹ & Bé',
}

def get_category(ma_code):
    """Trích tiền tố chữ cái từ Mã và map sang Ngành Hàng."""
    if pd.isna(ma_code):
        return 'Khác'
    prefix = ''
    for ch in str(ma_code).strip():
        if ch.isalpha():
            prefix += ch
        else:
            break
    return CATEGORY_MAP.get(prefix, 'Khác')

df_raw['Ngành Hàng'] = df_raw['Mã'].apply(get_category)
print(f"    → 'Ngành Hàng' added ({df_raw['Ngành Hàng'].nunique()} categories)")

# 3.2 Tỷ lệ lợi nhuận được tính lại (tránh chia cho 0)
df_raw['Tỷ lệ lợi nhuận'] = np.where(
    df_raw['Doanh Thu'] != 0,
    (df_raw['Lợi Nhuận'] / df_raw['Doanh Thu']).round(4),
    0
)
print(f"    → 'Tỷ lệ lợi nhuận' recalculated (avg: {df_raw['Tỷ lệ lợi nhuận'].mean():.2%})")

# 3.3 Doanh thu trên mỗi đơn vị sản phẩm
df_raw['Doanh Thu / Đơn vị'] = np.where(
    df_raw['SL Xuất'] != 0,
    (df_raw['Doanh Thu'] / df_raw['SL Xuất']).round(0),
    0
)
print(f"    → 'Doanh Thu / Đơn vị' calculated")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: KIỂM TRA DỮ LIỆU (Validation)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[4] Data validation...")

total_dt    = df_raw['Doanh Thu'].sum()
total_ln    = df_raw['Lợi Nhuận'].sum()
avg_margin  = total_ln / total_dt if total_dt != 0 else 0
n_orders    = len(df_raw)
n_customers = df_raw['Khách hàng'].nunique()
n_products  = df_raw['Hàng'].nunique()
aov         = total_dt / n_orders if n_orders != 0 else 0

print(f"    Doanh Thu   : {total_dt:>20,.0f} VNĐ")
print(f"    Lợi Nhuận   : {total_ln:>20,.0f} VNĐ")
print(f"    Margin      : {avg_margin:>19.2%}")
print(f"    Đơn hàng    : {n_orders:>20,}")
print(f"    Khách hàng  : {n_customers:>20,}")
print(f"    Sản phẩm    : {n_products:>20,}")
print(f"    AOV         : {aov:>20,.0f} VNĐ/đơn")

neg_rows = (df_raw['Lợi Nhuận'] < 0).sum()
if neg_rows > 0:
    print(f"    WARNING: {neg_rows} dòng có Lợi Nhuận âm")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: TẠO FILE OUTPUT — Copy toàn bộ sheets ngoại trừ Pivot
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n[5] Building output file: {OUTPUT_FILE}")
print(f"    Sheets to include : {COPY_SHEETS}")
print(f"    Sheets to SKIP    : {SKIP_SHEETS}")

# Chiến lược: dùng openpyxl để copy sheet nguyên vẹn (giữ format/chart/slicer)
# Riêng sheet Raw Data sẽ được GHI ĐÈ bằng df đã xử lý để có cột mới

# 5a. Copy toàn bộ file gốc trước (giữ dashboard + slicers)
import shutil
shutil.copy2(INPUT_FILE, OUTPUT_FILE)
print(f"    → Copied source to {OUTPUT_FILE}")

# 5b. Mở workbook bằng openpyxl và xóa các sheet Pivot
wb = openpyxl.load_workbook(OUTPUT_FILE)
for sheet_name in SKIP_SHEETS:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        print(f"    → Removed sheet: {sheet_name}")

# 5c. Ghi đè sheet Raw Data với dữ liệu đã được làm sạch + cột mới
if RAW_SHEET in wb.sheetnames:
    del wb[RAW_SHEET]

ws_raw_out = wb.create_sheet(RAW_SHEET)

# Style helpers
HEADER_FILL   = PatternFill('solid', fgColor='5B8C51')  # Mid green
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL      = PatternFill('solid', fgColor='EBF5E7')
THIN_BORDER   = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write header
for ci, col_name in enumerate(df_raw.columns, 1):
    cell = ws_raw_out.cell(row=1, column=ci, value=col_name)
    cell.font   = HEADER_FONT
    cell.fill   = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

# Write data rows
for ri, row in enumerate(df_raw.itertuples(index=False), 2):
    fill = ALT_FILL if ri % 2 == 0 else None
    for ci, value in enumerate(row, 1):
        # Convert NaT / NA to None for Excel compatibility
        if pd.isna(value) if not isinstance(value, str) else False:
            value = None
        cell = ws_raw_out.cell(row=ri, column=ci, value=value)
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

# Auto column width
for ci, col_name in enumerate(df_raw.columns, 1):
    max_len = max(len(str(col_name)), 12)
    ws_raw_out.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 35)

# Freeze top row
ws_raw_out.freeze_panes = 'A2'

# Auto-filter
ws_raw_out.auto_filter.ref = ws_raw_out.dimensions
print(f"    → Rewrote sheet '{RAW_SHEET}' with {len(df_raw)} rows + {len(df_raw.columns)} cols (incl. new cols)")

# 5d. Sắp xếp lại thứ tự sheet theo danh sách mong muốn
desired_order = [s for s in COPY_SHEETS if s in wb.sheetnames]
current_order = wb.sheetnames
for i, sheet_name in enumerate(desired_order):
    if sheet_name in wb.sheetnames:
        wb.move_sheet(sheet_name, offset=wb.sheetnames.index(sheet_name) - i)

wb.save(OUTPUT_FILE)
print(f"    → Saved {OUTPUT_FILE}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 6: BÁO CÁO KẾT QUẢ
# ─────────────────────────────────────────────────────────────────────────────
wb_check = openpyxl.load_workbook(OUTPUT_FILE, read_only=True)
final_sheets = wb_check.sheetnames
wb_check.close()

file_size_kb = os.path.getsize(OUTPUT_FILE) / 1024

print("\n" + "="*70)
print(f"  OUTPUT SUMMARY — {os.path.basename(OUTPUT_FILE)}")
print("="*70)
print(f"  File    : {OUTPUT_FILE}")
print(f"  Size    : {file_size_kb:,.1f} KB")
print(f"  Sheets  : {final_sheets}")
print(f"\n  Sheets REMOVED (Pivot):")
for s in SKIP_SHEETS:
    print(f"    ✗  {s}")
print(f"\n  New columns added to '{RAW_SHEET}':")
print(f"    +  Ngành Hàng              → phân loại sản phẩm từ Mã (12 nhóm)")
print(f"    +  Tỷ lệ lợi nhuận         → tính lại LN/DT (tránh chia 0)")
print(f"    +  Doanh Thu / Đơn vị      → DT ÷ SL Xuất")
print(f"\n  Tổng quan dữ liệu sau xử lý:")
print(f"    Doanh Thu   : {total_dt:>20,.0f} VNĐ")
print(f"    Lợi Nhuận   : {total_ln:>20,.0f} VNĐ ({avg_margin:.2%})")
print(f"    AOV         : {aov:>20,.0f} VNĐ/đơn")
print(f"    Đơn hàng    : {n_orders:>20,}")
print(f"    Khách hàng  : {n_customers:>20,}")
print(f"    Sản phẩm    : {n_products:>20,}")
print("="*70)
print(f"\nDone! Open: {OUTPUT_FILE}")
